import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.cell.cell import MergedCell
from functions import obtenerpath3

plantilla_excel = r'E:\Sharepoint\Pacífico Compañía de Seguros y Reaseguros\Analítica e Innovación en Auditoría - 01. Resultados Scripts\apps_innovacion\Auditron - GeneracionPEA\Result\plantilla_PEAS.xlsx'

SHEET_NAME = 'resumen'


# ---------------------------
# Helpers
# ---------------------------

def _normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns
          .str.strip()
          .str.lower()
          .str.replace(' ', '_')
          .str.replace('-', '_')
    )
    return df

def _etiqueta_item(idx: int) -> str:
    import string
    letters = string.ascii_lowercase
    s = ""
    n = idx
    while True:
        s = letters[n % 26] + s
        n = n // 26 - 1
        if n < 0:
            break
    return f"{s}."

def _get_plan_val(row_tuple) -> str:
    return getattr(row_tuple, 'plan_de_pruebas_din', getattr(row_tuple, 'plan_de_pruebas', ''))

def _anchor_of_merge(ws, row: int, col: int):
    """
    Si (row, col) cae dentro de un merged range, devuelve (min_row, min_col) de ese rango.
    Si no está en ningún merge, devuelve (row, col).
    """
    for mr in ws.merged_cells.ranges:
        if (mr.min_row <= row <= mr.max_row) and (mr.min_col <= col <= mr.max_col):
            return mr.min_row, mr.min_col
    return row, col

def _set_value_ws_rc(ws, row: int, col: int, value):
    """
    Escribe respetando merges: si (row,col) es MergedCell, redirige a la celda ancla.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        ar, ac = _anchor_of_merge(ws, row, col)
        ws.cell(row=ar, column=ac, value=value)
    else:
        cell.value = value

def _set_value_ws_coord(ws, coord: str, value):
    """
    Igual que _set_value_ws_rc pero recibiendo coordenada tipo 'C4'.
    """
    row, col = coordinate_to_tuple(coord)  # (row, col)
    _set_value_ws_rc(ws, row, col, value)


# ---------------------------
# Principal
# ---------------------------

def generar_informes(df_datos: pd.DataFrame,
                     df_procesada: pd.DataFrame,
                     codigo: str,
                     negocio: str):

    # Normaliza columnas
    df_datos = _normalizar_columnas(df_datos)
    df_procesada = _normalizar_columnas(df_procesada)

    # Merge de data
    df_merged = df_procesada.merge(df_datos, on='codigo_control', how='left', suffixes=('_din', '_fijo'))
    
    # Celdas fijas a poblar (pueden estar dentro de merges en la plantilla)
    celdas_fijas_map = {
        'C4':  'nombre_proyecto',
        'B6':  'codigo_riesgo',
        'B16': 'codigo_riesgo',
        'B12': 'objetivo',
        'B18': 'evento',
        'B22': 'codigo_control',
        'B24': 'descripcion_del_control',
    }

    # Map de clasificación -> nombre de tabla en la hoja
    clasif_to_table = {
        'prueba de cumplimiento': 'tabla_cumplimiento',
        'prueba sustantiva':      'tabla_sustantiva',
        'prueba multiproposito':  'tabla_multiproposito',
    }
    # Resuelve ruta final
    pathfinal = obtenerpath3(negocio, codigo)
    print(pathfinal)

    #Validar que si existe
    if os.path.exists("E:\\Sharepoint\\Pacífico Compañía de Seguros y Reaseguros\\Auditoria Interna - Evaluaciones-Documentos Salud\\2025\\PROGRAMADOS\\LAB - PRO - 003 - 2025\\Documentacion"):
        print ("existe2")
    else:
        print ("problemas2")
        
    # Por cada control, crea archivo y rellena
    for control, grupo in df_merged.groupby('codigo_control'):
        path_ctrl = os.path.join(pathfinal, str(control))
        os.makedirs(path_ctrl, exist_ok=True)

        destino = os.path.join(path_ctrl, f"PEA_{control}.xlsx")
        if not os.path.exists(destino):
            shutil.copy(plantilla_excel, destino)

        wb = load_workbook(destino)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        # Rellena celdas fijas (respetando merges)
        fila_fijo = grupo.iloc[0]
        for celda, col in celdas_fijas_map.items():
            _set_value_ws_coord(ws, celda, fila_fijo.get(col, ''))

        # Asegura comparaciones de texto homogéneas
        grupo = grupo.copy()
        grupo['clasificacion'] = grupo['clasificacion'].astype(str).str.lower()

        # Rellena tablas por clasificación
        for clasif, tabla_nombre in clasif_to_table.items():
            sub = grupo[grupo['clasificacion'] == clasif]
            if sub.empty:
                continue

            if tabla_nombre not in ws.tables:
                # La tabla no existe en la plantilla, saltar sin error
                continue

            tbl = ws.tables[tabla_nombre]
            min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
            data_start = min_row + 1  # la fila min_row suele ser el header de la tabla

            # Escribe filas (respetando merges si los hubiera)
            for idx, fila in enumerate(sub.itertuples(index=False)):
                row_dest = data_start + idx

                # Col 1: etiqueta a., b., c., ...
                _set_value_ws_rc(ws, row_dest, min_col, _etiqueta_item(idx))

                # Col 2: plan de pruebas (dinámico si existe, si no fijo)
                _set_value_ws_rc(ws, row_dest, min_col + 1, _get_plan_val(fila))

                # Resto de columnas: limpiar
                if max_col >= min_col + 2:
                    for c in range(min_col + 2, max_col + 1):
                        _set_value_ws_rc(ws, row_dest, c, None)

            # Ajusta el rango de la tabla para que abarque las filas nuevas
            # Si no hay filas, mantén al menos 1 fila de datos (Excel suele requerirlo)
            new_data_rows = max(len(sub), 1)
            new_max_row = data_start + new_data_rows - 1
            start_col_letter = get_column_letter(min_col)
            end_col_letter   = get_column_letter(max_col)
            tbl.ref = f"{start_col_letter}{min_row}:{end_col_letter}{new_max_row}"

        wb.save(destino)